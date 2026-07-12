from __future__ import annotations

import csv
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK_PATH = Path("src/assets/_DATA/GUNDAM_1.xlsx")
DATA_DIR = Path("src/assets/_DATA")
TABLES_DIR = DATA_DIR / "tables"
ANALYSIS_PATH = DATA_DIR / "workbook-analysis.json"
CATALOG_PATH = TABLES_DIR / "index.json"

EXPORTABLE_SHEETS = {
    "OLT": {
        "slug": "olt",
        "reason": "Catalogo maestro con 54 columnas bien definidas.",
    },
    "CORRESPONDENCIA": {
        "slug": "correspondencia",
        "reason": "Tabla de correspondencia OLT/SLOT/PTO/IP/CABLE/SPLITTER.",
    },
    "615": {
        "slug": "sheet_615",
        "reason": "Inventario de puertos y estados con encabezados completos.",
    },
    "PLANTILLA (TASK)": {
        "slug": "plantilla_task",
        "reason": "Plantillas CSTASK en formato tabular de asunto y descripcion.",
    },
    "RESPUESTAS (CSTASK)": {
        "slug": "respuestas_cstask",
        "reason": "Respuestas reutilizables CSTASK con estructura tabular simple.",
    },
    "Listas": {
        "slug": "listas",
        "reason": "Listado plano de perfiles reutilizables.",
    },
}

SKIPPED_SHEETS = {
    "COLTEL": "Mezcla JSON bruto, validaciones y layout visual; no es una tabla normalizada.",
    "PLANTILLA": "Contiene plantilla narrativa con una columna de JSON incrustado en el encabezado.",
    "NORMALIZACION": "Bloques de texto semi-estructurados, mas cercano a notas que a tabla.",
    "Matriz (N)": "Salida matricial orientada a comando y mensaje, sin encabezados reutilizables.",
    "Matriz (H)": "Salida matricial orientada a comando y mensaje, sin encabezados reutilizables.",
}


def normalize_header(value: Any, index: int, seen: dict[str, int]) -> str:
    if value is None or str(value).strip() == "":
        base_header = f"column_{index}"
    else:
        base_header = str(value).strip()

    duplicate_count = seen.get(base_header, 0)
    seen[base_header] = duplicate_count + 1

    if duplicate_count == 0:
        return base_header

    return f"{base_header}_{duplicate_count + 1}"


def normalize_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, date):
        return value.isoformat()

    return value


def sheet_headers(worksheet) -> list[str]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    seen: dict[str, int] = {}
    return [
        normalize_header(value, index + 1, seen)
        for index, value in enumerate(first_row)
    ]


def sheet_records(worksheet, headers: list[str]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        values = [normalize_cell(value) for value in row[: len(headers)]]

        if all(value in (None, "") for value in values):
            continue

        if len(values) < len(headers):
            values.extend([None] * (len(headers) - len(values)))

        records.append(dict(zip(headers, values, strict=False)))

    return records


def write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_csv(path: Path, headers: list[str], records: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(records)


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"No se encontro el archivo: {WORKBOOK_PATH}")

    TABLES_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    analysis: list[dict[str, Any]] = []
    catalog: list[dict[str, Any]] = []

    for worksheet in workbook.worksheets:
        headers = sheet_headers(worksheet)
        non_empty_headers = [header for header in headers if header.strip()]
        row_count = max(worksheet.max_row - 1, 0)
        is_exportable = worksheet.title in EXPORTABLE_SHEETS

        entry: dict[str, Any] = {
            "sheet": worksheet.title,
            "rows": worksheet.max_row,
            "dataRows": row_count,
            "columns": worksheet.max_column,
            "headerCount": len(non_empty_headers),
        }

        if is_exportable:
            config = EXPORTABLE_SHEETS[worksheet.title]
            records = sheet_records(worksheet, headers)
            json_path = TABLES_DIR / f"{config['slug']}.json"
            csv_path = TABLES_DIR / f"{config['slug']}.csv"

            write_json(json_path, records)
            write_csv(csv_path, headers, records)

            exported = {
                "sheet": worksheet.title,
                "slug": config["slug"],
                "rows": len(records),
                "columns": len(headers),
                "headers": headers,
                "jsonFile": str(json_path),
                "csvFile": str(csv_path),
                "reason": config["reason"],
            }
            catalog.append(exported)
            entry.update(
                {
                    "status": "exported",
                    "reason": config["reason"],
                    "output": {
                        "json": str(json_path),
                        "csv": str(csv_path),
                    },
                }
            )
        else:
            entry.update(
                {
                    "status": "skipped",
                    "reason": SKIPPED_SHEETS.get(
                        worksheet.title,
                        "No se clasifico como tabla reutilizable en esta primera extraccion.",
                    ),
                    "headerPreview": headers[:12],
                }
            )

        analysis.append(entry)

    write_json(
        CATALOG_PATH,
        {
            "sourceWorkbook": str(WORKBOOK_PATH),
            "tables": catalog,
        },
    )
    write_json(
        ANALYSIS_PATH,
        {
            "sourceWorkbook": str(WORKBOOK_PATH),
            "exportedTables": len(catalog),
            "sheets": analysis,
        },
    )


if __name__ == "__main__":
    main()
